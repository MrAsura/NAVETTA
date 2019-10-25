"""
Test suite containing functions for generating test results from TestInstances
"""

import openpyxl as xl
from openpyxl.utils import get_column_letter
import ast

import cfg
from .SummaryFactory import makeSummaries

__FILE_END = r".xlsm"
_KBS = r"kbs"
_KB = r"kb"
_TIME = r"time"
_PSNR = r"psnr"
_SCALE = r"scale"
_RES = r"results"
_QPS = r"qps"
_INAMES = r"inames"

__R_HEADER = ["Sequence","Layer"]
__R_HEADER_QP = "QP {}"
__R_KBS = ["Kb","Kb/s","Time (s)"]
__R_PSNR = "PSNR"
__R_PSNR_SUB = ["Y","U","V","AVG"]

__SEQ_FORMAT = "{} @ scale {}"
__PSNR_AVG = "=(6*{y}+{u}+{v})/8"

__C_AVG = r"=AVERAGE({})"
__SEQ_AVERAGE = r"Average"

_LID_TOT = -1

__LAYER2TEST_SEP = r"_layer"
__LAYER2TEST_FORMAT = r"{test}"+ __LAYER2TEST_SEP + "{lid}"

__COMBI_SEP = "+"
__LCOMBI_SEP = "_"

"""
Generate sheet layer string
"""
def makeSheetLayer(sheet,layer):
    return __LAYER2TEST_FORMAT.format(test=sheet,lid=layer)

"""
Parse sheet layer string
"""
def parseSheetLayer(string):
    return string.split(sep=__LAYER2TEST_SEP)

def parseCombiName(string):
    return string.split(sep=__COMBI_SEP)

def makeCombiName(combi):
    return __COMBI_SEP.join(combi)

def parseLayerCombiName(string):
    return string.split(sep=__LCOMBI_SEP)

def makeLayerCombiName(combi):
    return __LCOMBI_SEP.join(combi)


"""
Build test result dict
"""
def __resBuildFunc(results,seq,qp,lid,kbs,kb,time,psnr):
    
    if not seq in results:
        results[seq] = {}
    if not qp in results[seq]:
        results[seq][qp] = {}
    if not lid in results[seq][qp]:
        results[seq][qp][lid] = {}

    results[seq][qp][lid][_KBS] = kbs
    results[seq][qp][lid][_KB] = kb
    results[seq][qp][lid][_TIME] = time
    results[seq][qp][lid][_PSNR] = psnr

"""
Sort results qps and put them to ascending order by replacing them with order numbers
@return sorted dict and a dict containing the new to old qp key mapping
"""
def __sortQps(res):
    new_res = {}
    qp_map = {}
    for (seq,qps) in res.items():
        new_res[seq] = {}
        qp_map[seq] = {}

        sorted_qps = list(map(str, sorted([ast.literal_eval(qp) for qp in qps.keys()])))
        for (i,qp) in zip(range(len(sorted_qps)),sorted_qps):
            new_res[seq][i] = qps[qp]
            qp_map[seq][i] = qp

    return (new_res,qp_map)

"""
Parse test results
@return a dict with parsed results in the form of result[<test_name>] = {__RES: {<seq>: {<qp>: <lid>:{__KBS,__KB,__TIME,__PSNR}}}, __SCALE, __QPS, __INAMES }
"""
def __parseTestResults(tests):
    results = {}
    for test in tests:
        main_res = test.getResults(__resBuildFunc,l_tot=_LID_TOT)
        (main_res,qp_names) = __sortQps(main_res)
        results[test._test_name] = {_RES: main_res, _SCALE: str(test._input_layer_scales), _QPS: qp_names, _INAMES: [__SEQ_AVERAGE,] + test.getInputNames()}
    return results

"""
Combine psnr and kbs values
"""
def __combiValues(vals):
   
    res = {_RES:{},_SCALE:'',_QPS:{}, _INAMES:vals[0][_INAMES]}
    #Init structure
    for (seq,qps) in vals[0][_RES].items():
        res[_RES][seq] = {}
        res[_QPS][seq] = {}
        for (qp,lids) in qps.items():
            res[_RES][seq][qp] = {}
            res[_QPS][seq][qp] = ""
            for (lid,val) in lids.items():
                res[_RES][seq][qp][lid] = {}
                res[_RES][seq][qp][lid][_KBS] = 0
                res[_RES][seq][qp][lid][_KB] = 0
                res[_RES][seq][qp][lid][_TIME] = 0
                res[_RES][seq][qp][lid][_PSNR] = (0,0,0)

    numv = len(vals)

    scales = []
    for item in vals:
        scales.append(item[_SCALE])
        for (seq,qps) in item[_RES].items():
            for (qp,lids) in qps.items():
                for (lid,val) in lids.items():
                    res[_RES][seq][qp][lid][_KBS] += val[_KBS]
                    res[_RES][seq][qp][lid][_KB] += val[_KB]
                    res[_RES][seq][qp][lid][_TIME] += val[_TIME]
                    res[_RES][seq][qp][lid][_PSNR] = tuple(map(lambda x,y: float(y) + float(x)/float(numv), val[_PSNR], res[_RES][seq][qp][lid][_PSNR]))
                res[_QPS][seq][qp] = makeCombiName([res[_QPS][seq][qp],item[_QPS][seq][qp]]) if len(res[_QPS][seq][qp]) > 0 else item[_QPS][seq][qp]
    res[_SCALE] = makeCombiName(scales)
        
    return res

"""
Combine psnr and kbs values in a layered fashion
"""
def __layerCombiValues(vals):
   
    res = {_RES:{},_SCALE:'',_QPS:{}, _INAMES:vals[0][_INAMES]}
    #Init structure
    for (seq,qps) in vals[0][_RES].items():
        res[_RES][seq] = {}
        res[_QPS][seq] = {}
        for qp in qps.keys():
            res[_RES][seq][qp] = {}
            res[_QPS][seq][qp] = "()"
            for lid in range(len(vals)):
                res[_RES][seq][qp][lid] = {}
                res[_RES][seq][qp][lid][_KBS] = 0
                res[_RES][seq][qp][lid][_KB] = 0
                res[_RES][seq][qp][lid][_TIME] = 0
                res[_RES][seq][qp][lid][_PSNR] = (0,0,0)
            res[_RES][seq][qp][_LID_TOT] = {}
            res[_RES][seq][qp][_LID_TOT][_KBS] = 0
            res[_RES][seq][qp][_LID_TOT][_KB] = 0
            res[_RES][seq][qp][_LID_TOT][_TIME] = 0
            res[_RES][seq][qp][_LID_TOT][_PSNR] = (0,0,0)

    numv = len(vals)

    scales = []
    for item in vals:
        scales.append(item[_SCALE])

    for (seq,qps) in res[_RES].items():
        for (qp,lids) in qps.items():
            for (lid,val) in zip(range(numv),vals):
                lids[lid][_KBS] = val[_RES][seq][qp][_LID_TOT][_KBS]
                lids[lid][_KB] = val[_RES][seq][qp][_LID_TOT][_KB]
                lids[lid][_TIME] = val[_RES][seq][qp][_LID_TOT][_TIME]
                lids[lid][_PSNR] = val[_RES][seq][qp][_LID_TOT][_PSNR]
                lids[_LID_TOT][_KBS] += val[_RES][seq][qp][_LID_TOT][_KBS]#/numv #Take the average
                lids[_LID_TOT][_KB] += val[_RES][seq][qp][_LID_TOT][_KB]
                lids[_LID_TOT][_TIME] += val[_RES][seq][qp][_LID_TOT][_TIME]
                lids[_LID_TOT][_PSNR] = tuple(map(lambda x,y: float(y) + float(x)/float(numv), val[_RES][seq][qp][_LID_TOT][_PSNR], lids[_LID_TOT][_PSNR]))

                res[_QPS][seq][qp] = str( ast.literal_eval(res[_QPS][seq][qp]) + ast.literal_eval(val[_QPS][seq][qp]) )

    res[_SCALE] = makeLayerCombiName(scales)
        
    return res

"""
Combine test results to form new tests
@param combi: list of results to combine  (or list of lists). 
@param layer_combi: list of results to combine. layer_combi[i] will be the ith layer.
"""
def __combiTestResults(results, combi, layer_combi):
    res = results.copy()

    for set in combi:
        vals = []
        for item in set:
            vals.append(results[item])
        cname = makeCombiName(set)
        res[cname] = __combiValues(vals)

    for set in layer_combi:
        vals = []
        for item in set:
            vals.append(results[item])
        cname = makeLayerCombiName(set)
        res[cname] = __layerCombiValues(vals)

    return res
    

"""
Write results for a single test/sheet
@return positions of relevant cells as res_ref[<seq>][<lid>] = {__KB, __KBS,__PSNR, __TIME}
"""
def __writeSheet(sheet,data,scale,qp_names,order=None):
    # Write header
    for col in range(len(__R_HEADER)):
        sheet.cell(row = 1, column = col+1).value = __R_HEADER[col]
        sheet.merge_cells(start_column=col+1,start_row=1,end_column=col+1,end_row=3)
        sheet.cell(row=1,column=col+1).alignment = xl.styles.Alignment(horizontal='center')

    sheet.column_dimensions['A'].width = 50

    qp_cols = {}
    seq_rows = {}
    res_ref = {}

    order = order if order else list(data.keys())

    # Write stat row
    for (qp,item) in sorted(tuple(data.values())[0].items()):
        qp_cols[qp] = sheet.max_column+1
        for val in __R_KBS:
            sheet.cell(row=2,column=sheet.max_column+1).value = val
            sheet.merge_cells(start_column=sheet.max_column,start_row=2,end_column=sheet.max_column,end_row=3)
            sheet.cell(row=2,column=sheet.max_column).alignment = xl.styles.Alignment(horizontal='center')

        for val in __R_PSNR_SUB:
            sheet.cell(row=3,column=sheet.max_column+1).value = val
        sheet.cell(row=2,column=sheet.max_column-len(__R_PSNR_SUB)+1).value = __R_PSNR
        sheet.merge_cells(start_column=sheet.max_column-len(__R_PSNR_SUB)+1,start_row=2,end_column=sheet.max_column,end_row=2)
        sheet.cell(row=2,column=sheet.max_column-len(__R_PSNR_SUB)+1).alignment = xl.styles.Alignment(horizontal='center')
        
        sheet.cell(row=1,column=qp_cols[qp]).value = __R_HEADER_QP.format(tuple(qp_names.values())[0][qp])
        sheet.merge_cells(start_column=qp_cols[qp],start_row=1,end_column=sheet.max_column,end_row=1)
        sheet.cell(row=1,column=qp_cols[qp]).alignment = xl.styles.Alignment(horizontal='center')

    # Write sequence column
    #for (seq,res) in data.items():
    layer_r = range(len(tuple(tuple(data.values())[0].values())[0].keys())-1)
    for seq in order:
        sheet.cell(row=sheet.max_row+1,column=1).value = __SEQ_FORMAT.format(seq,scale) if seq is not __SEQ_AVERAGE else __SEQ_AVERAGE 
        seq_rows[seq] = sheet.max_row

        res_ref[seq] = {}
            
        #Set Layers
        sheet.cell(row=seq_rows[seq],column=2).value = _LID_TOT
        res_ref[seq][_LID_TOT] = {_KB:[], _KBS:[],_PSNR:[],_TIME:[]}
        for lid in layer_r:
            sheet.cell(row=seq_rows[seq]+lid+1,column=2).value = lid
            res_ref[seq][lid] = {_KB:[], _KBS:[],_PSNR:[],_TIME:[]}
    
    # Set actual data
    #for (seq,qps) in data.items():
    for seq in order:
        if seq in data:
            qps = data[seq]
            for (qp,item) in sorted(qps.items()):
                for (lid,val) in item.items():
                    r = seq_rows[seq] + lid + 1
                    if lid == _LID_TOT:
                        r = seq_rows[seq]
                    c_kb = qp_cols[qp]
                    c_kbs = c_kb + 1
                    c_time = c_kbs + 1
                    c_psnr = c_time + len(__R_PSNR_SUB)

                    sheet.cell(row=r,column=c_kb).value = val[_KB]
                    sheet.cell(row=r,column=c_kbs).value = val[_KBS]
                    sheet.cell(row=r,column=c_time).value = val[_TIME]

                    for i in range(len(__R_PSNR_SUB)-1):
                        sheet.cell(row=r,column=c_psnr-i-1).value = float(val[_PSNR][-i-1])
                    sheet.cell(row=r,column=c_psnr).value = __PSNR_AVG.format(y = get_column_letter(c_psnr-3) + str(r),
                                                                              u = get_column_letter(c_psnr-2) + str(r),
                                                                              v = get_column_letter(c_psnr-1) + str(r))

                    res_ref[seq][lid][_KB].append(get_column_letter(c_kb) + str(r))
                    res_ref[seq][lid][_KBS].append(get_column_letter(c_kbs) + str(r))
                    res_ref[seq][lid][_TIME].append(get_column_letter(c_time) + str(r))
                    res_ref[seq][lid][_PSNR].append(get_column_letter(c_psnr) + str(r))
        else:
            # Set Average data
            for c_kb in sorted(qp_cols.values()):
                for lid in res_ref[__SEQ_AVERAGE]:
                    r = seq_rows[__SEQ_AVERAGE] + lid + 1
                    if lid == _LID_TOT:
                        r = seq_rows[__SEQ_AVERAGE]
                    c_kbs = c_kb + 1
                    c_time = c_kbs + 1
                    c_psnr = c_time + len(__R_PSNR_SUB)

                    kb_rows = []
                    kbs_rows = []
                    time_rows = []
                    for (seq,row) in seq_rows.items():
                        if seq == __SEQ_AVERAGE:
                            continue
                        kb_rows.append(get_column_letter(c_kb)+str(row+lid+1))
                        kbs_rows.append(get_column_letter(c_kbs)+str(row+lid+1))
                        time_rows.append(get_column_letter(c_time)+str(row+lid+1))
                    sheet.cell(row=r,column=c_kb).value = __C_AVG.format(','.join(kb_rows))
                    sheet.cell(row=r,column=c_kbs).value = __C_AVG.format(','.join(kbs_rows))
                    sheet.cell(row=r,column=c_time).value = __C_AVG.format(','.join(time_rows))

                    for i in range(len(__R_PSNR_SUB)-1):
                        psnr_rows = []
                        for (seq,row) in seq_rows.items():
                            if seq == __SEQ_AVERAGE:
                                continue
                            psnr_rows.append(get_column_letter(c_psnr-i-1)+str(row+lid+1))
                        sheet.cell(row=r,column=c_psnr-i-1).value = __C_AVG.format(','.join(psnr_rows))
            
                    sheet.cell(row=r,column=c_psnr).value = __PSNR_AVG.format(y = get_column_letter(c_psnr-3) + str(r),
                                                                              u = get_column_letter(c_psnr-2) + str(r),
                                                                              v = get_column_letter(c_psnr-1) + str(r))

                    res_ref[__SEQ_AVERAGE][lid][_KB].append(get_column_letter(c_kb) + str(r))
                    res_ref[__SEQ_AVERAGE][lid][_KBS].append(get_column_letter(c_kbs) + str(r))
                    res_ref[__SEQ_AVERAGE][lid][_TIME].append(get_column_letter(c_time) + str(r))
                    res_ref[__SEQ_AVERAGE][lid][_PSNR].append(get_column_letter(c_psnr) + str(r))


    sheet.freeze_panes = 'A4'

    return res_ref


"""
Write results to workbook. Assume wb has a Summary and Result Sheet.
"""
def __writeResults(wb,results,summary_defs):
    res_pos = {}
    
    # Write test results
    for (test,res) in sorted(results.items()):
        n_sheet = wb.create_sheet(title=test,index=0)
        res_pos[test] = __writeSheet(n_sheet,res[_RES],res[_SCALE],res[_QPS],res[_INAMES])

    #write summary sheets
    makeSummaries(wb, res_pos, *summary_defs,
                  order = res[_INAMES])

"""
Run given tests and write results to a exel file
@param combi: Give a list of test names that are combined into one test
@param layer_combi: Given tests are combined as like they were layers
@param layers: A dict with test names as keys containing a list of layers to include in summary
@param s2_base: Test name of the s2 summary that should be the base of the comparison
"""
def runTests( tests, outname, *summary_defs, combi = [], layer_combi = [], input_res = False):
    print('Start running tests...')
    nt = 1
    for test in tests:
        #print("Running test {}...".format(test._test_name))
        print_out = "[{}/{}] ".format(nt,len(tests))
        print(print_out, end='\r')
        test.run(print_out, input_res)
        nt += 1
    print('Tests complete.')
    print('Writing results to file {}...'.format(cfg.results + outname + __FILE_END))
    res = __parseTestResults(tests)
    res = __combiTestResults(res,combi,layer_combi)

    wb = xl.load_workbook(cfg.exel_template,keep_vba=True)
    wb.remove(wb.active) #Remove un-used sheet
    __writeResults(wb,res,summary_defs)
    wb.save(cfg.results + outname + __FILE_END)
    print('Done.')
