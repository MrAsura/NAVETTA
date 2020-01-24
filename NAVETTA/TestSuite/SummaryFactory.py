"""
Module for defining and processing summary sheets
"""

import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import ScatterChart, Reference, Series
from enum import Enum, auto
from typing import Dict, Tuple, Union, List, Callable, Iterable

#Define summary types used as keys in definitions
class SummaryType(Enum):
    NONE = 0
    BDBRM = auto()
    ANCHOR = auto()
    CURVE = auto()

__TYPE_NAMES = {
    SummaryType.BDBRM: "BDBRMatrix",
    SummaryType.ANCHOR: "Anchor_list",
    SummaryType.CURVE: "Curve_chart"
    }

# The container definition for summary types
__TYPE = "Type" #Type of the summary that is defined
__NAME = "Name" #Name for the summary.  Will be used as the sheet name
__DEFINITION = "definition" #Should contain one of the summary type definitions below
"""
__DEFINITION: [dt_BDBRM | dt_ANCHOR | ...]
"""
dt_BASE = {__TYPE: SummaryType.NONE, __NAME: str(), __DEFINITION: dict()}

def create_summary_definition(type: SummaryType, definition: dict, name: str = None) -> dict:
    new_def = dt_BASE.copy()
    new_def[__TYPE] = type
    new_def[__DEFINITION] = definition
    name = name.strip() if name.strip() else __TYPE_NAMES[type]
    new_def[__NAME] = name
    return new_def

######### Define definition templates for each summary type ############

"""
__LAYERS:{<test_name>:<tuple of layers (lid) to include>}
"""
__LAYERS = "layers"
__WRITE_BITS = "write_bits"
__WRITE_BDBR = "write_bdbr"
__WRITE_PSNR = "write_psnr"
__WRITE_TIME = "write_time"
dt_BDBRM = {__LAYERS:{}, __WRITE_BDBR: True, __WRITE_BITS: True, __WRITE_PSNR: True, __WRITE_TIME: True}

"""
for creating the BDBRMatrix definition
"""
def create_BDBRMatrix_definition(layers: Dict[str, Tuple[int]], write_bdbr: bool, write_bits: bool, write_psnr: bool, write_time: bool, name: str = "") -> dict:
    definition = dt_BDBRM.copy()
    definition[__LAYERS] = layers
    definition[__WRITE_BDBR] = write_bdbr
    definition[__WRITE_BITS] = write_bits
    definition[__WRITE_PSNR] = write_psnr
    definition[__WRITE_TIME] = write_time
    return create_summary_definition(SummaryType.BDBRM, definition, name)

"""
__????:dt_ANCHOR_SUB|None
"""
__BDBR = "bdbr"
__BITS = "bits"
__PSNR = "psnr"
__TIME = "time"
dt_ANCHOR = {__BDBR:{}, __BITS:{}, __PSNR:{}, __TIME:{}}
"""
dt_ANCHOR_SUB = {<test_name>:(<anchor_test_name>|None,...)}
"""

AnchorSubType = Dict[str, Tuple[Union[str,None],...]]

"""
Create anchor list definition. Pass in sub definitions for each data type that should be included
"""
def create_AnchorList_definition(bdbr_def: AnchorSubType, bits_def: AnchorSubType, psnr_def: AnchorSubType, time_def: AnchorSubType, name: str = "") -> dict:
    definition = dt_ANCHOR.copy()
    definition[__BDBR] = create_AnchorSub_definition(bdbr_def)
    definition[__BITS] = create_AnchorSub_definition(bits_def)
    definition[__PSNR] = create_AnchorSub_definition(psnr_def)
    definition[__TIME] = create_AnchorSub_definition(time_def)
    return create_summary_definition(SummaryType.ANCHOR, definition, name)

"""
Sub anchor list definition. Takes in a list of test anchor pairs with optional layer ids ((<name>,#lid):(<anchor>,#lid))
"""
def create_AnchorSub_definition(definition: Union[None, Tuple[Union[str, Tuple[str, int]], Union[str, Tuple[str,int]]], Dict[Union[str, Tuple[str,int]], Union[str, Tuple[str,int]]]]) -> AnchorSubType:
    from .TestSuite import makeSheetLayer, _LID_TOT
    def processTestName(tn):
        if not tn:
            return None
        if isinstance(tn, str):
            return tn
        if len(tn) == 2 and isinstance(tn, tuple):
            if isinstance(tn[0], str) and isinstance(tn[1], int):
                return makeSheetLayer(tn[0], tn[1]) if tn[1] != _LID_TOT else tn[0]
        return tuple(processTestName(test) for test in tn)

    if not definition:
        return None
    if isinstance(definition, tuple):
        return {processTestName(definition[0]): processTestName(definition[1])}
    if isinstance(definition, dict):
        definition = definition.items()
    return {processTestName(name): processTestName(val) for (name, val) in definition if val}

"""
__TESTS: ((<name>,#lid) | <name>,...)  
__CHARTS:((<X_Data_type>,<Y_Data_type>),...)
?_Data_type: 'psnr' | 'rate' | 'time' 
"""
__TESTS = "tests"
__CHARTS = "charts"
__RATE = "rate"
dt_CURVE = {__TESTS:(), __CHARTS:()}

"""
Create Curve Chart definition.
"""
def create_CurveChart_definition(tests: Iterable[Union[str, Tuple[str, int]]], charts: Iterable[Tuple[str,str]], name: str = ""):
    def processTestName(tn):
        if not tn:
            return None
        if isinstance(tn, str):
            return tn
        if len(tn) == 2 and isinstance(tn, tuple):
            if isinstance(tn[0], str) and isinstance(tn[1], int):
                return makeSheetLayer(tn[0], tn[1]) if tn[1] != _LID_TOT else tn[0]
        return tuple(processTestName(test) for test in tn)

    definition = dt_CURVE.copy()
    definition[__TESTS] = processTestName(tests)
    chart_data_types = [__PSNR, __RATE, __TIME]
    definition[__CHARTS] = [c for c in charts if c[0] in chart_data_types and c[1] in chart_data_types]
    
    return create_summary_definition(SummaryType.CURVE, definition, name)

DataRefType = Dict[str,Dict[str,Dict[int,dict]]] #DataRef[<test_name>][<seq>][<lid>] = {__KB, __KBS,__PSNR, __TIME}
SummaryRefType = Dict[str,Dict[str,dict]] #SummaryRef[<test_name>][<seq>] = {__KB, __KBS,__PSNR, __TIME}

"""
Function for creating summary sheets to the given workbook for the given data
@param wb: work book where sheets will be added
@param data_refs: references to the data sheets and data cell positions used for the summary in the form data_refs[<test_name>][<seq>][<lid>] = {__KB, __KBS,__PSNR, __TIME}
@param BDBRMatrix: dict containing the parameters for the BDBRMatrix summary type
"""
def makeSummaries(wb: xl.Workbook, data_refs: DataRefType, *definitions: List[dict], order: List[str] = None) -> None:
    for definition in definitions:
        if SummaryType.BDBRM == definition[__TYPE]:
            makeBDBRMatrix(wb, data_refs, order, definition)
        elif SummaryType.ANCHOR == definition[__TYPE]:
            makeAnchorList(wb, data_refs, order, definition)
        elif SummaryType.CURVE == definition[__TYPE]:
            makeCurveChart(wb, data_refs, order, definition)
        else:
            print("Not a valid summary type.")


"""
Return a name for a sheet that does not yet exist based on the given name (return name if it does not exist yet else name#)
"""
def __get_new_sheetname(wb: xl.Workbook, name: str) -> str:
    if name in wb.sheetnames:
        #If name already exists, add a number to the end and try again
        postfix = 2
        while name + str(postfix) in wb.sheetnames:
            postfix += 1
        name += str(postfix)
    return name

def makeBDBRMatrix(wb: xl.Workbook, data_refs: DataRefType, order: List[str], definition: dict) -> None:
    name = __get_new_sheetname(wb, definition[__NAME])
    bdbrm_sheet = wb.create_sheet(name)
    expanded_refs = __makeSummary(data_refs, definition[__DEFINITION][__LAYERS])
    __writeBDBRMatrix(bdbrm_sheet, expanded_refs, order, **definition[__DEFINITION])
    wb.active = wb.index(wb[name])

def makeAnchorList(wb: xl.Workbook, data_refs: DataRefType, order: List[str], definition: dict) -> None:
    name = __get_new_sheetname(wb, definition[__NAME])
    anchor_sheet = wb.create_sheet(name)
    expanded_refs = __makeSummary(data_refs)
    __writeAnchorList(anchor_sheet, expanded_refs, order, **definition[__DEFINITION])
    wb.active = wb.index(wb[name])

def makeCurveChart(wb: xl.Workbook, data_refs: DataRefType, order: List[str], definition: dict) -> None:
    name = __get_new_sheetname(wb, definition[__NAME])
    curve_sheet = wb.create_sheet(name)
    expanded_refs = __makeSummary(data_refs)
    __writeCurveChart(curve_sheet, expanded_refs, order, **definition[__DEFINITION])
    wb.active = wb.index(wb[name])

"""
Transform res_pos into summary test structure making test layers into their own tests
@return dict of the form res[<test_name>][<seq>] = {__KB, __KBS,__PSNR, __TIME}
"""
def __makeSummary(res_pos: DataRefType, layers: dict = {}) -> SummaryRefType:
    from .TestSuite import _LID_TOT, makeSheetLayer
    res = {}
    for (test,item) in res_pos.items():
        res[test] = {}
        for (seq,vals) in item.items():
            test_in_layers = test in layers
            if not test_in_layers:
                res[test][seq] = vals[_LID_TOT]
                if len(vals.keys()) <= 2: #If only 2 lids, it should mean the total layer and other layer are the same
                    continue
            for (lid,val) in vals.items():
                if lid == _LID_TOT and not test_in_layers:
                    continue
                if test_in_layers and lid not in layers[test]:
                    continue
                nn = makeSheetLayer(test,lid)
                if lid == _LID_TOT or len(vals.keys()) <= 2:
                    nn = test
                if nn in res:
                    res[nn][seq] = val
                else:
                    res[nn] = {seq:val}

    return res

"""
Switch the first and second level keys of the dict
"""
def __flip_dict(old_d: dict) -> dict:
    new_d = {val_key: {} for val in old_d.values() for val_key in val.keys()}
    for (key, val) in old_d.items():
        for (val_key, sub_val) in val.items():
            new_d[val_key][key] = sub_val
    return new_d


def getMaxLength( string_list: list ) -> str:
    return max((len(x) for x in string_list))


__S_BDRATE_FORMAT = "=bdrate({},{},{},{},{},{},{},{},{},{},{},{},{},{},{},{})"
__S_BIT_FORMAT = "=AVERAGE({},{},{},{})/AVERAGE({},{},{},{})"
__S_TIME_FORMAT = "=AVERAGE({},{},{},{})/AVERAGE({},{},{},{})"
__S_PSNR_FORMAT = "=AVERAGE({},{},{},{})-AVERAGE({},{},{},{})"
__SR_FORMAT = r"'{sheet}'!{cell}"

__S_BIT_ABS_FORMAT = "=AVERAGE({},{},{},{})"
__S_TIME_ABS_FORMAT = "=AVERAGE({},{},{},{})"
__S_PSNR_ABS_FORMAT = "=AVERAGE({},{},{},{})"

#######################################
# AnchorList summary type definitions #
#######################################
__AL_HEADER = r"Result anchor list summary"
__AL_SEQ_FORMAT = r"{}"
__AL_BD_HEADER = r"BDBR results"
__AL_B_HEADER = r"Bit results"
__AL_PSNR_HEADER = r"PSNR results"
__AL_TIME_HEADER = r"Time results"
__AL_TEST = r"Test:"
__AL_ANCHOR = r"Sequences \ Anchor:"
#__AL_SEQ = r"Sequences"

"""
Handle writing the anchor list structure
"""

def __writeAnchorList(sheet: Worksheet, data_refs: SummaryRefType, order: List[str] = None, *, bdbr: AnchorSubType, bits: AnchorSubType, psnr: AnchorSubType, time: AnchorSubType, **other: dict) -> None:
    from .TestSuite import _PSNR, _KBS, _KB, _TIME
    seq_ref = __flip_dict(data_refs) # transform data_refs to seq_ref[<seq>][<test_name>] order
    order = order if order else list(seq_ref.keys())

    # Each sequence is one line so generate one columns for each test in each gategory based on the given definitions
    sheet.cell(row = 1, column = 1).value = __AL_HEADER 
    static_row = 1
    
    for seq in order:

        header_offset = 0
        row = sheet.max_row + 1
        #col = 0 #sheet.max_column + 1
        
        if 'first_row' not in locals():
            first_row = row + 1

        # write bdrate tests
        if bdbr:
            if 'bdcol' not in locals():
                bdcol = sheet.max_column + 3
                sheet.cell(row = static_row, column = bdcol - 1).value = __AL_BD_HEADER
                sheet.cell(row = static_row + 1, column = bdcol - 1).value = __AL_TEST
                sheet.cell(row = static_row + 2, column = bdcol - 1).value = __AL_ANCHOR
                #sheet.cell(row = static_row + 3, column = bdcol - 1).value = __AL_SEQ
                bdcol_end = __writeAnchorListHeader(sheet, bdbr, static_row + 1, bdcol, False)
                header_offset = static_row + 1

            __writeAnchorListData(sheet, seq_ref[seq], bdbr, row + header_offset, bdcol,
                                     data_func = lambda data, test: data[test][_KBS] + data[test][_PSNR],
                                     data_format = __S_BDRATE_FORMAT,
                                     number_format = '0.00%')
                # Write sequence
            sheet.cell(row = row + header_offset, column = bdcol - 1).value = __AL_SEQ_FORMAT.format(seq)
                                   

        # write bit tests
        if bits:
            if 'bcol' not in locals():
                bcol = sheet.max_column + 3
                sheet.cell(row = static_row, column = bcol - 1).value = __AL_B_HEADER
                sheet.cell(row = static_row + 1, column = bcol - 1).value = __AL_TEST
                sheet.cell(row = static_row + 2, column = bcol - 1).value = __AL_ANCHOR
                #sheet.cell(row = static_row + 3, column = bcol - 1).value = __AL_SEQ
                bcol_end = __writeAnchorListHeader(sheet, bits, static_row + 1, bcol)
                header_offset = static_row + 1

            __writeAnchorListData(sheet, seq_ref[seq], bits, row + header_offset, bcol,
                                     data_func = lambda data, test: data[test][_KB],
                                     data_format = __S_BIT_FORMAT,
                                     abs_format = __S_BIT_ABS_FORMAT)
                # Write sequence
            sheet.cell(row = row + header_offset, column = bcol - 1).value = __AL_SEQ_FORMAT.format(seq)

        # write psnr tests
        if psnr:
            if 'pcol' not in locals():
                pcol = sheet.max_column + 3
                sheet.cell(row = static_row, column = pcol - 1).value = __AL_PSNR_HEADER
                sheet.cell(row = static_row + 1, column = pcol - 1).value = __AL_TEST
                sheet.cell(row = static_row + 2, column = pcol - 1).value = __AL_ANCHOR
                #sheet.cell(row = static_row + 3, column = pcol - 1).value = __AL_SEQ
                pcol_end = __writeAnchorListHeader(sheet, psnr, static_row + 1, pcol)
                header_offset = static_row + 1
            
            __writeAnchorListData(sheet, seq_ref[seq], psnr, row + header_offset, pcol,
                                     data_func = lambda data, test: data[test][_PSNR],
                                     data_format = __S_PSNR_FORMAT,
                                     abs_format = __S_PSNR_ABS_FORMAT)
                # Write sequence
            sheet.cell(row = row + header_offset, column = pcol - 1).value = __AL_SEQ_FORMAT.format(seq)

        # write time matrix
        if time:
            if 'tcol' not in locals():
                tcol = sheet.max_column + 3
                sheet.cell(row = static_row, column = tcol - 1).value = __AL_TIME_HEADER
                sheet.cell(row = static_row + 1, column = tcol - 1).value = __AL_TEST
                sheet.cell(row = static_row + 2, column = tcol - 1).value = __AL_ANCHOR
                #sheet.cell(row = static_row + 3, column = tcol - 1).value = __AL_SEQ
                tcol_end = __writeAnchorListHeader(sheet, time, static_row + 1, tcol)
                header_offset = static_row + 1
            
            __writeAnchorListData(sheet, seq_ref[seq], time, row + header_offset, tcol,
                                     data_func = lambda data, test: data[test][_TIME],
                                     data_format = __S_TIME_FORMAT,
                                     abs_format = __S_TIME_ABS_FORMAT)

            # Write sequence
            sheet.cell(row = row + header_offset, column = tcol - 1).value = __AL_SEQ_FORMAT.format(seq)


    # Make columns wider
    for column in range(sheet.max_column):
        sheet.column_dimensions[get_column_letter(column+1)].width = getMaxLength(list(data_refs.keys()) + list(order))

    #Add conditional formatting
    form_ranges = []
    color_rules = []
    #BDRATE
    if bdbr:
        form_ranges.append("{}:{}".format(get_column_letter(bdcol)+str(first_row),get_column_letter(bdcol_end)+str(row)))
        color_rules.append(ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                                          end_type='percentile', end_value=10, end_color='F8696B' ))
    if bits:
        form_ranges.append("{}:{}".format(get_column_letter(bcol)+str(first_row),get_column_letter(bcol_end)+str(row)))
        color_rules.append(ColorScaleRule(start_type='min', start_color='4F81BD',
                                          mid_type='num', mid_value=1, mid_color='FFFFFF',
                                          end_type='percentile', end_value=80, end_color='F8696B' ))
    if psnr:
        form_ranges.append("{}:{}".format(get_column_letter(pcol)+str(first_row),get_column_letter(pcol_end)+str(row)))
        color_rules.append(ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                                          end_type='percentile', end_value=10, end_color='F8696B' ))
    if time:
        form_ranges.append("{}:{}".format(get_column_letter(tcol)+str(first_row),get_column_letter(tcol_end)+str(row)))
        color_rules.append(ColorScaleRule(end_type='min', start_color='9BDE55',#'63BE7B',
                                          mid_type='num', mid_value=1, mid_color='FFFFFF',
                                          start_type='percentile', end_value=80, end_color='00BBEF'))

    for (f_range, c_rule) in zip(form_ranges, color_rules):
        sheet.conditional_formatting.add(f_range, c_rule)


def __writeAnchorListHeader(sheet: Worksheet, sub_def: AnchorSubType, row: int, col: int, allow_none: bool = True) -> None:
    #Write horizontal headers/test names
    tmp_col = col
    for (test,anchors) in sub_def.items():
        for anchor in anchors:
            if not allow_none and not anchor:
                continue
            sheet.cell(row = row, column = tmp_col).value = test
            sheet.cell(row = row + 1, column = tmp_col).value = anchor
            tmp_col += 1
    return tmp_col - 1


def __writeAnchorListData(sheet: Worksheet, ref: Dict[str, dict], sub_def: AnchorSubType, row: int, col: int, *, data_func: Callable[[Dict[str, dict], str], Union[float,int]], data_format: str, number_format: Union[None, str] = None, number_style: str = 'Percent', abs_format: Union[None, str] = None, abs_style: str = 'Comma') -> None:
    from .TestSuite import parseSheetLayer
    #final_r = row+len(data.keys())
    #final_c = col+len(data.keys())
    c = col
    for (test, anchors) in sub_def.items():
        for anchor in anchors:
            value_format = data_format
            value_style = number_style
            anchor_res = []
            if not anchor:
                if abs_format:
                    value_format = abs_format
                    if abs_style:
                        value_style = abs_style
                else:
                    continue
            else:
                anchor_res =[__SR_FORMAT.format(sheet=parseSheetLayer(anchor)[0],cell=cl) for cl in data_func(ref, anchor)]
            test_res =[__SR_FORMAT.format(sheet=parseSheetLayer(test)[0],cell=cl) for cl in data_func(ref, test)]
            sheet.cell(row = row, column = c).value = value_format.format(*(anchor_res + test_res))
            sheet.cell(row = row, column = c).style = value_style
            if number_format:
                sheet.cell(row = row, column = c).number_format = number_format
            sheet.cell(row= row, column= c).alignment = xl.styles.Alignment(horizontal='center')
            c += 1
    # Set conditional coloring
    #form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    #sheet.conditional_formatting.add(form_range, color_scale_rule)


#######################################
# BDBRMatrix summary type definitions #
#######################################

__S_SEQ_HEADER = "Sequence {} results"
__S_BIT_HEADER = r"Bit comparisons"
__S_PSNR_HEADER = r"PSNR comparisons (dB)"
__S_TIME_HEADER = r"Encoding time comparisons"
__S_HEADER = "Result summary matrix (bdrate, bit, PSNR, Time comparisons)"


"""
Handle writing the BDBRMatrix summary sheet
"""
def __writeBDBRMatrix(sheet: Worksheet, data_refs: SummaryRefType, order: List[str] = None, *, write_bdbr: bool, write_bits:bool, write_psnr: bool, write_time: bool, **other: dict):
    from .TestSuite import _PSNR, _KBS, _KB, _TIME
    
    seq_ref = __flip_dict(data_refs) # transform data_refs to seq_ref[<seq>][<test_name>] order
    order = order if order else list(seq_ref.keys())

    #print(seq_ref)
    # For each sequence generate the comparison matrix
    sheet.cell(row = 1, column = 1).value = __S_HEADER 
    #for (seq,ref) in sorted(seq_ref.items()):
    for seq in order:
        ref = seq_ref[seq]
        tests = sorted(ref.keys())
        
        row = sheet.max_row + 2
        brow = row
        prow = row
        trow = row
        col = 1 #sheet.max_column + 1

        sheet.cell(row = row, column = col).value = __S_SEQ_HEADER.format(seq) #Write sequence header
        
        # write bdrate matrix
        if write_bdbr:  

            sheet.merge_cells(start_column=col,start_row=row,end_column=col+len(tests),end_row=row)
            (row, col) = __writeSummaryMatrixHeader(sheet, tests, row+1, col)
            __writeSummaryDataMatrix(sheet, ref, row, col,
                                     data_func = lambda data, test: data[test][_KBS] + data[test][_PSNR],
                                     data_format = __S_BDRATE_FORMAT,
                                     number_format = '0.00%',
                                     color_scale_rule = ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=10, end_color='F8696B' ))
                                   

        # write bit matrix
        if write_bits:
            if 'bcol' not in locals():
                bcol = sheet.max_column + 2
            sheet.cell(row = brow, column = bcol).value = __S_BIT_HEADER
            sheet.merge_cells(start_column=bcol,start_row=brow,end_column=bcol+len(tests),end_row=brow)
            (brow, col) = __writeSummaryMatrixHeader(sheet, tests, brow+1, bcol)
            __writeSummaryDataMatrix(sheet, ref, brow, colb,
                                     data_func = lambda data, test: data[test][_KB],
                                     data_format = __S_BIT_FORMAT,
                                     color_scale_rule = ColorScaleRule(start_type='min', start_color='4F81BD',
                                                                       mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=80, end_color='F8696B' ))

        # write psnr matrix
        if write_psnr:
            if 'pcol' not in locals():
                pcol = sheet.max_column + 2
            sheet.cell(row = prow, column = pcol).value = __S_PSNR_HEADER
            sheet.merge_cells(start_column=pcol,start_row=prow,end_column=pcol+len(tests),end_row=prow)
            (prow, col) = __writeSummaryMatrixHeader(sheet, tests, prow+1, pcol)
            __writeSummaryDataMatrix(sheet, ref, prow, colb,
                                     data_func = lambda data, test: data[test][_PSNR],
                                     data_format = __S_PSNR_FORMAT,
                                     number_style = 'Comma',
                                     def_val = 0,
                                     color_scale_rule = ColorScaleRule(start_type='percentile', start_value=90, start_color='63BE7B',
                                                                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=10, end_color='F8696B' ))

        # write time matrix
        if write_time:
            if 'tcol' not in locals():
                tcol = sheet.max_column + 2
            sheet.cell(row = trow, column = tcol).value = __S_TIME_HEADER
            sheet.merge_cells(start_column=tcol,start_row=trow,end_column=tcol+len(tests),end_row=trow)
            (trow, col) = __writeSummaryMatrixHeader(sheet, tests, trow+1, tcol)
            __writeSummaryDataMatrix(sheet, ref, trow, col,
                                     data_func = lambda data, test: data[test][_TIME],
                                     data_format = __S_TIME_FORMAT,
                                     color_scale_rule = ColorScaleRule(start_type='min', start_color='9BDE55',#'63BE7B',
                                                                       mid_type='num', mid_value=1, mid_color='FFFFFF',
                                                                       end_type='percentile', end_value=80, end_color='00BBEF'))

    # Make columns wider
    for col in range(sheet.max_column):
        sheet.column_dimensions[get_column_letter(col+1)].width = getMaxLength(list(data_refs.keys()))


"""
Write summary base header
@return row and col of start of data field
"""
def __writeSummaryMatrixHeader(sheet: Worksheet, tests: List[str], row: int, col: int) -> Tuple[int,int]:
    d_row = row + 1
    d_col = col + 1
    #Write horizontal headers/test names
    tmp_col = col + 1
    for test in tests:
        sheet.cell(row = row, column = tmp_col).value = test
        tmp_col += 1

    #Write vertical 
    tmp_row = row + 1
    for test in tests:
        sheet.cell(row = tmp_row, column = col).value = test
        tmp_row += 1

    return d_row, d_col

"""
Write summary matrix data array
"""
def __writeSummaryDataMatrix(sheet: Worksheet, data: Dict[str, dict], row: int, col: int, *, data_func: Callable[[Dict[str, dict], str], Union[float,int]], data_format: str, number_format: Union[str, None] = None, number_style: str = 'Percent', def_val: str = '-', color_scale_rule: Rule):
    from .TestSuite import parseSheetLayer
    test_col = col - 1
    test_row = row - 1
    final_r = row+len(data.keys())
    final_c = col+len(data.keys())
    for r in range(row,row+len(data.keys())):
        for c in range(col,col+len(data.keys())):
            t2 = sheet.cell(row = r, column = test_col).value
            t1 = sheet.cell(row = test_row, column = c).value
            if t1 == t2:
                sheet.cell(row = r, column = c).value = def_val
            else:
                r1 =[__SR_FORMAT.format(sheet=parseSheetLayer(t1)[0],cell=cl) for cl in data_func(data, t1)]
                r2 =[__SR_FORMAT.format(sheet=parseSheetLayer(t2)[0],cell=cl) for cl in data_func(data, t2)]
                sheet.cell(row = r, column = c).value = data_format.format(*(r1+r2))
                sheet.cell(row = r, column = c).style = number_style
                if number_format:
                    sheet.cell(row = r, column = c).number_format = number_format
            sheet.cell(row=r,column=c).alignment = xl.styles.Alignment(horizontal='center')
    # Set conditional coloring
    form_range = "{}:{}".format(get_column_letter(col)+str(row),get_column_letter(final_c)+str(final_r))
    sheet.conditional_formatting.add(form_range, color_scale_rule)



#######################################
# curve_chart summary type definitions #
#######################################

def __writeCurveChart(sheet: Worksheet, data_refs: SummaryRefType, order: List[str] = None, *, tests: Iterable[Union[str]], charts: Iterable[Tuple[str,str]], **other: dict) -> None:
    refs = __writeCurveChartData(sheet, tests, order, data_refs)
    __writeCharts(sheet, tests, order, charts, refs)

# Write data used by charts and return per test Reference ranges
def __writeCurveChartData(sheet: Worksheet, tests: Iterable[str], order: Iterable[str], data: Dict[str, dict]) -> dict:
    from .TestSuite import _PSNR, _KBS, _TIME, parseSheetLayer
    to_write_data = []
    out_ref = {}
    col = 2
    ref_len = 5

    def toRefFunc(sheet, cells):
        return ["=" + __SR_FORMAT.format(sheet=parseSheetLayer(sheet)[0], cell = cell) for cell in cells]

    for seq in order:
        to_write_data.append([seq,])
        out_ref[seq] = {}
        to_write_data.append(["Data Type", "Test", "Data Point 1", "Data Point 2", "Data Point 3", "Data Point 4"])

        for test in tests:
            out_ref[seq][test] = {}

            to_write_data.append([_KBS, test, *toRefFunc(test, data[test][seq][_KBS])]) 
            to_write_data.append([_PSNR, test, *toRefFunc(test, data[test][seq][_PSNR])]) 
            to_write_data.append([_TIME, test, *toRefFunc(test, data[test][seq][_TIME])])

            row = len(to_write_data)

            out_ref[seq][test][__RATE] = Reference(sheet, min_col=col, max_col=col + ref_len, min_row = row - 2)
            out_ref[seq][test][__PSNR] = Reference(sheet, min_col=col, max_col=col + ref_len, min_row = row - 1)
            out_ref[seq][test][__TIME] = Reference(sheet, min_col=col, max_col=col + ref_len, min_row = row)

    for row in to_write_data:
        sheet.append(row)

    # hide chart data
    sheet.column_dimensions.group('A', 'F', hidden = True)

    return out_ref

# Create chart objects based on the given Reference ranges and 
def __writeCharts(sheet: Worksheet, tests: Iterable[str], order: Iterable[str], charts: Iterable[Tuple[str,str]], data: Dict[str, dict]) -> None:
    row = 1
    for seq in order:
        col = 0
        for (typeX, typeY) in charts:
            chart = ScatterChart(scatterStyle = 'lineMarker')
            chart.title = seq
            chart.x_axis.title = typeX
            chart.y_axis.title = typeY
            chart.visible_cells_only = False
            for test in tests:
                rX = data[seq][test][typeX]
                rY = data[seq][test][typeY]
                series = Series(rY, Reference(sheet, min_col = rX.min_col + 1, max_col = rX.max_col, min_row = rX.min_row), title_from_data = True)
                series.marker.symbol = 'auto'
                chart.series.append(series)
            sheet.add_chart(chart, chr(ord('G') + col) + str(row))
            col += 9
        row += 15
